"""
Excel export service for class rosters and individual child reports.

Uses openpyxl to generate .xlsx files with formatted worksheets.
"""

import io
from datetime import datetime
from typing import Dict, Any, List, Optional
from sqlalchemy import select, desc
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Child, AbilityAssessment, Report


async def export_class_roster(
    db: AsyncSession,
    class_name: str,
) -> bytes:
    """
    Generate an Excel class roster with latest assessment scores.

    Sheet 1: 班级花名册 — child name, age_group, latest per-dimension scores, levels
    Sheet 2: 维度汇总 — per-dimension average, level distribution
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("需要安装 openpyxl: pip install openpyxl")

    # Fetch children in class
    result = await db.execute(
        select(Child).where(Child.class_name == class_name).order_by(Child.name)
    )
    children = result.scalars().all()

    if not children:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "空班级"
        ws["A1"] = "该班级没有幼儿"
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # Fetch latest assessments per child
    child_assessments = {}
    all_dimensions = set()
    for child in children:
        ar = await db.execute(
            select(AbilityAssessment)
            .where(AbilityAssessment.child_id == child.id)
            .order_by(desc(AbilityAssessment.assessed_at))
        )
        assessments = ar.scalars().all()
        # Take latest per dimension
        latest = {}
        for a in assessments:
            if a.dimension not in latest:
                latest[a.dimension] = a
        child_assessments[child.id] = latest
        all_dimensions.update(latest.keys())

    dims_sorted = sorted(all_dimensions)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Roster ──
    ws1 = wb.active
    ws1.title = "班级花名册"

    # Header
    headers = ["姓名", "年龄段", "班级"]
    for dim in dims_sorted:
        headers.append(f"{dim}分数")
        headers.append(f"{dim}水平")
    ws1.append(headers)

    # Bold header
    for cell in ws1[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Data rows
    for child in children:
        row = [
            child.name,
            child.age_group.value if hasattr(child.age_group, 'value') else str(child.age_group),
            child.class_name or "",
        ]
        latest = child_assessments.get(child.id, {})
        for dim in dims_sorted:
            a = latest.get(dim)
            row.append(a.score if a else "")
            row.append(a.level.value if a and a.level else "")
        ws1.append(row)

    # Auto-width
    for col in ws1.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # ── Sheet 2: Dimension Summary ──
    ws2 = wb.create_sheet("维度汇总")
    ws2.append(["维度", "平均分", "最高分", "最低分", "评分人数"])

    for cell in ws2[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    dim_scores = {dim: [] for dim in dims_sorted}
    for child_id, latest in child_assessments.items():
        for dim, a in latest.items():
            dim_scores[dim].append(a.score)

    for dim in dims_sorted:
        scores = dim_scores[dim]
        if scores:
            ws2.append([dim, round(sum(scores) / len(scores), 1), max(scores), min(scores), len(scores)])
        else:
            ws2.append([dim, "", "", "", 0])

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


async def export_child_report(
    db: AsyncSession,
    child_id: int,
) -> bytes:
    """
    Generate an Excel file with a single child's complete assessment history.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("需要安装 openpyxl: pip install openpyxl")

    child = await db.get(Child, child_id)
    if not child:
        raise ValueError(f"Child not found: {child_id}")

    ar = await db.execute(
        select(AbilityAssessment)
        .where(AbilityAssessment.child_id == child_id)
        .order_by(desc(AbilityAssessment.assessed_at))
    )
    assessments = ar.scalars().all()

    rr = await db.execute(
        select(Report)
        .where(Report.child_id == child_id)
        .order_by(desc(Report.generated_at))
    )
    reports = rr.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{child.name}的评估报告"

    # Header info
    ws.append(["姓名", child.name])
    ws.append(["年龄段", child.age_group.value if hasattr(child.age_group, 'value') else str(child.age_group)])
    ws.append(["班级", child.class_name or ""])
    ws.append(["分析次数", len(reports)])
    ws.append([])

    # Assessment history
    ws.append(["评估维度", "分数", "水平", "PCK阶段", "错误模式", "建议", "评估日期"])
    for cell in ws[6]:
        cell.font = openpyxl.styles.Font(bold=True)

    for a in assessments:
        error_str = ", ".join(a.error_patterns) if a.error_patterns else ""
        ws.append([
            a.dimension,
            a.score,
            a.level.value if hasattr(a.level, 'value') else str(a.level),
            a.pck_stage or "",
            error_str,
            a.recommendations or "",
            a.assessed_at.strftime("%Y-%m-%d %H:%M") if a.assessed_at else "",
        ])

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
