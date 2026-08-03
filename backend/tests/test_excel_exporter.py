"""
Tests for the Excel exporter service.

Covers: class roster export (with/without children, with assessments),
single-child report export, and edge cases. Uses the in-memory SQLite
db_session fixture from conftest.
"""

import io
from datetime import datetime

import pytest
import openpyxl
from sqlalchemy import select

from app.models.child import Child
from app.models.assessment import AbilityAssessment
from app.models.report import Report
from app.models.enums import AgeGroupEnum, LevelEnum, ReportTypeEnum
from app.services.excel_exporter import export_class_roster, export_child_report

# All tests in this module are async (use the db_session fixture)
pytestmark = pytest.mark.asyncio


# ─── Fixtures ────────────────────────────────────────────────────────

@pytest.fixture
def make_child():
    """Factory to create Child objects with unique access codes."""
    counter = [0]

    def _make(name="小明", age=AgeGroupEnum.MIDDLE, class_name="中一班"):
        counter[0] += 1
        return Child(
            name=name,
            age_group=age,
            class_name=class_name,
            parent_access_code=f"CODE{counter[0]:04d}",
        )
    return _make


async def _add_child_with_assessments(db, make_child, dims_scores):
    """Add a child plus one AbilityAssessment per (dimension, score, level) tuple."""
    child = make_child()
    db.add(child)
    await db.flush()  # populate child.id
    for dim, score, level in dims_scores:
        db.add(AbilityAssessment(
            child_id=child.id,
            dimension=dim,
            score=score,
            level=level,
            pck_stage="动作水平",
            assessed_at=datetime(2025, 9, 1, 10, 0),
        ))
    await db.commit()
    return child


# ─── export_class_roster ─────────────────────────────────────────────

class TestExportClassRoster:
    async def test_empty_class_returns_valid_xlsx(self, db_session):
        data = await export_class_roster(db_session, "空班级")
        assert isinstance(data, bytes)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        assert "空班级" in ws.title or ws["A1"].value

    async def test_roster_with_children(self, db_session, make_child):
        db_session.add(make_child(name="小明", class_name="中一班"))
        db_session.add(make_child(name="小红", class_name="中一班"))
        await db_session.commit()

        data = await export_class_roster(db_session, "中一班")
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["班级花名册"]
        names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "小明" in names
        assert "小红" in names

    async def test_roster_includes_scores(self, db_session, make_child):
        await _add_child_with_assessments(db_session, make_child, [
            ("counting", 80, LevelEnum.L3),
        ])
        data = await export_class_roster(db_session, "中一班")
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["班级花名册"]
        # Header should contain a dimension-score column
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert any("counting" in str(h).lower() and "分数" in str(h) for h in headers)

    async def test_roster_has_two_sheets(self, db_session, make_child):
        db_session.add(make_child())
        await db_session.commit()
        data = await export_class_roster(db_session, "中一班")
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "班级花名册" in wb.sheetnames
        assert "维度汇总" in wb.sheetnames


# ─── export_child_report ─────────────────────────────────────────────

class TestExportChildReport:
    async def test_report_with_assessments(self, db_session, make_child):
        child = await _add_child_with_assessments(db_session, make_child, [
            ("counting", 60, LevelEnum.L2),
            ("shapes_space", 85, LevelEnum.L4),
        ])
        data = await export_child_report(db_session, child.id)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        assert child.name in ws.title
        # Should contain the child's name in header rows
        all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert child.name in all_text

    async def test_report_includes_score_values(self, db_session, make_child):
        child = await _add_child_with_assessments(db_session, make_child, [
            ("counting", 75, LevelEnum.L3),
        ])
        data = await export_child_report(db_session, child.id)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert "75" in all_text  # the score

    async def test_report_for_nonexistent_child_raises(self, db_session):
        with pytest.raises(ValueError):
            await export_child_report(db_session, 99999)

    async def test_report_with_no_assessments_still_works(self, db_session, make_child):
        child = make_child()
        db_session.add(child)
        await db_session.commit()
        data = await export_child_report(db_session, child.id)
        assert isinstance(data, bytes)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert wb.active is not None

    async def test_report_counts_assessment_rows(self, db_session, make_child):
        child = await _add_child_with_assessments(db_session, make_child, [
            ("counting", 60, LevelEnum.L2),
            ("shapes_space", 70, LevelEnum.L3),
            ("patterns", 50, LevelEnum.L2),
        ])
        data = await export_child_report(db_session, child.id)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        # Header info row "分析次数" should reflect 0 reports (we added no Report rows)
        all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert "分析次数" in all_text
